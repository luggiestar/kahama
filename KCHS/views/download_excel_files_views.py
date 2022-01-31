import xlrd
import xlwt
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.defaultfilters import upper
from xlsxwriter import Workbook

from ..models import *


def download_course_assessment_excel(request, assessment, course):
    get_group_assessment = GroupAssessment.objects.get(id=assessment)
    get_semester = AcademicSemester.objects.get(is_active=True)

    get_course = ProgrammeCourseStructure.objects.get(course__code=course)
    # content-type of response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    template_name = str(get_course.course.code) + "_" + str(get_group_assessment.item) + "_" + str(
        get_group_assessment.category) + "_" + str(
        get_semester.academic_year.year) + "_result"
    name = f"{template_name}.xlsx"
    response['Content-Disposition'] = 'attachment; filename=' + name
    # book = Workbook(response, {'in_memory': True})
    # sheet = book.add_worksheet('sheet1')
    #

    # creating workbook
    wb = Workbook(response, {'in_memory': True})

    # adding sheet
    ws = wb.add_worksheet("sheet1")
    ws.set_column('B:B', 18)
    ws.set_column('C:C', 30)
    ws.set_column('D:D', 23)
    ws.set_column('E:E', 11)
    ws.set_column('F:F', 19)

    # Sheet header, first row
    row_num = 0
    # Add a bold format to use to highlight cells.
    bold = wb.add_format({'bold': 1, 'font_color': 'red', 'font_name': 'Cambria'})

    # bold.set_font_name('Times New Roman')
    # font_style = xlwt.XFStyle()
    # headers are bold
    # font_style.font.bold = True

    # column header names, you can use your own headers here
    columns = ['S/N', 'Registration#', 'Full Name', 'Programme', 'Course', 'Assessment', 'Marks']

    # write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    # Sheet body, remaining rows
    # font_style = xlwt.XFStyle()

    # get your data, from database or from a text file...
    data = SemesterAssessment.objects.filter(programme_course=get_course, assessment_group=get_group_assessment,
                                             academic_semester=get_semester)  # dummy method to fetch data.
    for my_row in data:
        get_name = f"{my_row.registration.student.user.first_name} {my_row.registration.student.user.middle_name} {my_row.registration.student.user.last_name} "
        get_course = f"{my_row.programme_course.course}"
        get_assessment = f"{get_group_assessment.item} ({get_group_assessment.category})"
        get_reg_number = f"{my_row.registration.student.user}"
        get_program = f"{my_row.registration.student.programme}"
        row_num = row_num + 1
        ws.write(row_num, 0, row_num)
        ws.write(row_num, 1, get_reg_number)
        ws.write(row_num, 2, get_name)
        ws.write(row_num, 3, get_program)

        ws.write(row_num, 4, get_course)
        ws.write(row_num, 5, get_assessment)
        ws.write(row_num, 6, my_row.marks, bold)

    wb.close()
    return response


def download_course_result_excel(request, course):
    # get_group_assessment = GroupAssessment.objects.get(id=assessment)
    get_semester = AcademicSemester.objects.get(is_active=True)

    get_course = ProgrammeCourseStructure.objects.get(course__code=course)

    # content-type of response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    template_name = str(get_course.course.code) + "_" + str(get_semester.semester) + "_" + str(
        get_semester.academic_year.year) + "_results"
    name = f"{template_name}.xlsx"
    response['Content-Disposition'] = 'attachment; filename=' + name
    # book = Workbook(response, {'in_memory': True})
    # sheet = book.add_worksheet('sheet1')
    #

    # creating workbook
    wb = Workbook(response, {'in_memory': True})

    # adding sheet
    ws = wb.add_worksheet("sheet1")
    ws.set_column('A:A', 4)
    ws.set_column('B:B', 18)
    ws.set_column('C:C', 30)
    ws.set_column('D:D', 20)
    ws.set_column('E:E', 6)
    ws.set_column('F:F', 6)
    ws.set_column('G:G', 10)

    ws.set_column('H:H', 8)
    ws.set_column('I:I', 13)

    # Sheet header, first row
    row_num = 0
    # Add a bold format to use to highlight cells.
    bold = wb.add_format({'bold': 1, 'font_color': 'red', 'font_name': 'Cambria'})

    # bold.set_font_name('Times New Roman')
    # font_style = xlwt.XFStyle()
    # headers are bold
    # font_style.font.bold = True

    # column header names, you can use your own headers here
    columns = ['S/N', 'Registration#', 'Full Name', 'Programme', 'CA', 'ES', 'TOTAL', 'GRADE', 'REMARK']

    # write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    # Sheet body, remaining rows
    # font_style = xlwt.XFStyle()

    # get your data, from database or from a text file...
    data = SemesterResult.objects.filter(programme_course=get_course,
                                         academic_semester=get_semester)  # dummy method to fetch data.
    for my_row in data:
        get_name = f"{my_row.registration.student.user.first_name} {my_row.registration.student.user.middle_name} {my_row.registration.student.user.last_name} "
        get_course = f"{my_row.programme_course.course}"
        # get_assessment = f"{get_group_assessment.item} ({get_group_assessment.category})"
        get_reg_number = f"{my_row.registration.student.user}"
        get_program = f"{my_row.registration.student.programme}"
        row_num = row_num + 1
        ws.write(row_num, 0, row_num)
        ws.write(row_num, 1, get_reg_number)
        ws.write(row_num, 2, get_name)
        ws.write(row_num, 3, get_program)

        ws.write(row_num, 4, my_row.ca)
        ws.write(row_num, 5, my_row.es)
        ws.write(row_num, 6, my_row.total, bold)
        ws.write(row_num, 7, my_row.grade)
        ws.write(row_num, 8, my_row.remark, bold)

    wb.close()
    return response


# def student_course_assessment_templateb(request, assessment, course):
#     get_group_assessment = GroupAssessment.objects.get(id=assessment)
#     get_semester = AcademicSemester.objects.get(is_active=True)
#
#     get_course = ProgrammeCourseStructure.objects.get(course__code=course)
#     # content-type of response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#
#     # decide file name
#     template_name = str(get_course.course.code) + "_" + str(get_group_assessment.item) + "_" + "_" + str(get_semester)
#     name = f"{template_name}.xls"
#     response['Content-Disposition'] = 'attachment; filename=' + name
#
#     # creating workbook
#     wb = xlwt.Workbook(encoding='utf-8')
#
#     # adding sheet
#     ws = wb.add_sheet("sheet1")
#
#     # Sheet header, first row
#     row_num = 0
#
#     font_style = xlwt.XFStyle()
#     # headers are bold
#     font_style.font.bold = True
#
#     # column header names, you can use your own headers here
#     columns = ['S/N', 'Full Name', 'Course', 'Assessment', 'Marks']
#
#     # write column headers in sheet
#     for col_num in range(len(columns)):
#         ws.write(row_num, col_num, columns[col_num], font_style)
#
#     # Sheet body, remaining rows
#     font_style = xlwt.XFStyle()
#
#     # get your data, from database or from a text file...
#
#     data = Registration.objects.filter(student__programme=get_course.programme,
#                                        semester=get_semester).exclude(
#         id__in=SemesterAssessment.objects.filter(programme_course=get_course, assessment_group=get_group_assessment,
#                                              academic_semester=get_semester).values(
#             'registration__id'))  # dummy method to fetch data.
#     get_course = f"{get_course.course}"
#     get_assessment = f"{get_group_assessment.item}"
#     for my_row in data:
#         get_name = f"{my_row.student.user.first_name} {my_row.student.user.middle_name} {my_row.student.user.last_name} "
#
#         row_num = row_num + 1
#         ws.write(row_num, 0, row_num, font_style)
#         ws.write(row_num, 1, get_name, font_style)
#         ws.write(row_num, 2, get_course, font_style)
#         ws.write(row_num, 3, get_assessment, font_style)
#         ws.write(row_num, 4, "", font_style)
#
#     wb.save(response)
#     return response


def student_course_assessment_template(request, assessment, course):
    get_group_assessment = GroupAssessment.objects.get(id=assessment)
    get_semester = AcademicSemester.objects.get(is_active=True)

    get_course = ProgrammeCourseStructure.objects.get(course__code=course)
    # content-type of response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    template_name = str(get_course.course.code) + "_" + str(get_group_assessment.item) + "_" + str(
        get_group_assessment.category) + "_template"
    name = f"{template_name}.xlsx"
    response['Content-Disposition'] = 'attachment; filename=' + name

    # book = Workbook(response, {'in_memory': True})
    # sheet = book.add_worksheet('sheet1')
    #

    # creating workbook
    wb = Workbook(response, {'in_memory': True})

    # adding sheet
    ws = wb.add_worksheet("sheet1")
    ws.set_column('B:B', 17)
    ws.set_column('C:C', 23)
    ws.set_column('D:D', 23)
    ws.set_column('E:E', 11)
    ws.set_column('F:F', 19)

    # Sheet header, first row
    row_num = 0
    # Add a bold format to use to highlight cells.
    bold = wb.add_format({'bold': 1, 'font_color': 'red', 'font_name': 'Cambria'})

    # bold.set_font_name('Times New Roman')
    # font_style = xlwt.XFStyle()
    # headers are bold
    # font_style.font.bold = True

    # column header names, you can use your own headers here
    columns = ['S/N', 'Registration#', 'Full Name', 'Programme', 'Course', 'Assessment', 'Marks']

    # write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    # Sheet body, remaining rows
    # font_style = xlwt.XFStyle()

    # get your data, from database or from a text file...

    data = Registration.objects.filter(student__programme=get_course.programme,
                                       semester=get_semester).exclude(
        id__in=SemesterAssessment.objects.filter(programme_course=get_course, assessment_group=get_group_assessment,
                                                 academic_semester=get_semester).values(
            'registration__id'))  # dummy method to fetch data.
    get_course = f"{get_course.course}"
    get_assessment = f"{get_group_assessment.item} ({get_group_assessment.category})"
    for my_row in data:
        get_name = f"{my_row.student.user.first_name} {my_row.student.user.middle_name} {my_row.student.user.last_name} "
        get_reg_number = f"{my_row.student.user}"
        get_program = f"{my_row.student.programme}"

        row_num = row_num + 1
        ws.write(row_num, 0, row_num)
        ws.write(row_num, 1, get_reg_number)
        ws.write(row_num, 2, get_name)
        ws.write(row_num, 3, get_program)
        ws.write(row_num, 4, get_course)
        ws.write(row_num, 5, get_assessment)
        ws.write(row_num, 6, "", bold)
    wb.close()
    return response


def student_entry_template(request, ):
    get_semester = AcademicSemester.objects.get(is_active=True)

    # content-type of response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    template_name = str(get_semester.semester) + "_" + str(get_semester.academic_year.year) + "_entry_template"
    name = f"{template_name}.xlsx"
    response['Content-Disposition'] = 'attachment; filename=' + name

    # book = Workbook(response, {'in_memory': True})
    # sheet = book.add_worksheet('sheet1')
    #

    # creating workbook
    wb = Workbook(response, {'in_memory': True})

    # adding sheet
    ws = wb.add_worksheet("sheet1")
    ws.set_column('A:A', 15)
    ws.set_column('B:B', 15)
    ws.set_column('C:C', 15)
    ws.set_column('D:D', 15)
    ws.set_column('E:E', 10)
    ws.set_column('F:F', 15)
    ws.set_column('G:G', 28)
    ws.set_column('H:H', 15)

    # Sheet header, first row
    row_num = 0
    # Add a bold format to use to highlight cells.
    bold = wb.add_format({'bold': 1, 'font_color': 'red', 'font_name': 'Cambria'})

    # bold.set_font_name('Times New Roman')
    # font_style = xlwt.XFStyle()
    # headers are bold
    # font_style.font.bold = True

    # column header names, you can use your own headers here
    columns = ['index Number', 'First_ Name', 'Middle Name', 'Last Name', 'sex', 'phone', 'programme', 'Entry Level']

    # write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    # Sheet body, remaining rows
    # font_style = xlwt.XFStyle()

    # get your data, from database or from a text file...

    wb.close()
    return response


def delete_course_assessment_result_data(request, assessment, course):
    get_group_assessment = GroupAssessment.objects.get(id=assessment)
    get_semester = AcademicSemester.objects.get(is_active=True)

    get_course = ProgrammeCourseStructure.objects.get(course__code=course)
    print(get_group_assessment.item)

    SemesterAssessment.objects.filter(programme_course=get_course,
                                      assessment_group__item=get_group_assessment.item,
                                      academic_semester=get_semester).delete()
    if get_group_assessment.category == "ES":
        get_result = SemesterResult.objects.filter(programme_course=get_course,
                                                   academic_semester=get_semester)
        for i in get_result:
            i.es = decimal.Decimal(0)
            i.total = decimal.Decimal(0)
            i.grade = "-"
            i.remark = "-"
            i.save()

    return redirect('KCHS:course_assessment_result', assessment=get_group_assessment.id, course=get_course.course)


from django.shortcuts import render
import openpyxl


def upload_course_assessment(request, assessment, course):
    get_group_assessment = GroupAssessment.objects.get(id=assessment)
    get_semester = AcademicSemester.objects.get(is_active=True)

    get_course = ProgrammeCourseStructure.objects.get(course__code=course)
    if "GET" == request.method:
        return redirect('KCHS:course_assessment_result', assessment=get_group_assessment.id, course=get_course.course)
    else:
        excel_file = request.FILES['excel_file']

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]
        # print(worksheet)

        # iterating over the rows and
        # getting value from each cell in row
        # try:

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            get_student = Registration.objects.get(
                student__user__username=worksheet.cell(row=rowno, column=2).value, semester=get_semester)

            # get_weight = (worksheet.cell(row=rowno, column=7).value / decimal.Decimal(100)) * decimal.Decimal(
            #     get_group_assessment.weight)

            save_data = SemesterAssessment.objects.update_or_create(
                registration=get_student,
                academic_semester=get_semester,
                assessment_group=get_group_assessment,
                programme_course=get_course,
                marks=worksheet.cell(row=rowno, column=7).value
                # weight=get_weight

            )

        messages.success(request, f"Successfully Uploaded")

        # except:
        #     messages.error(request, f"Failed, Please  Upload only required Data")
        #     return redirect('KCHS:course_assessment_result', assessment=get_group_assessment.id,
        #                     course=get_course.course)

        return redirect('KCHS:course_assessment_result', assessment=get_group_assessment.id, course=get_course.course)


def id_generator(size=3, chars=string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def upload_student_entry(request):
    get_semester = AcademicSemester.objects.get(is_active=True)

    if "GET" == request.method:
        return redirect('KCHS:course_assessment_result')
    else:
        excel_file = request.FILES['excel_file']

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]
        # print(worksheet)


        # iterating over the rows and
        # getting value from each cell in row
        # try:

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), start=2):
            get_program = Programme.objects.get(name="CLINICAL MEDICINE")
            get_level = Level.objects.get(name=worksheet.cell(row=rowno, column=8).value)
            get_username = f"KACHS-{id_generator()}-{get_semester.academic_year.year}"
            create_email = f"{get_username}@kachs.ac.tz"
            get_phone = f"0{worksheet.cell(row=rowno, column=6).value}"

            # get_weight = (worksheet.cell(row=rowno, column=7).value / decimal.Decimal(100)) * decimal.Decimal(
            #     get_group_assessment.weight)

            save_user = User(
                username=get_username,
                first_name=worksheet.cell(row=rowno, column=2).value,
                middle_name=worksheet.cell(row=rowno, column=3).value,
                last_name=worksheet.cell(row=rowno, column=4).value,
                sex=worksheet.cell(row=rowno, column=5).value,
                email=create_email,
                phone=get_phone,
                password=make_password(worksheet.cell(row=rowno, column=4).value),
                title="student"

            )
            save_user.save()
            save_student = Student(
                user=User.objects.get(username=get_username),
                programme=get_program,
                entry_level=get_level

            )
            save_student.save()

        messages.success(request, f"Successfully Uploaded")

        # except:
        #     messages.error(request, f"Failed, Please  Upload only required Data")
        #     return redirect('KCHS:course_assessment_result', assessment=get_group_assessment.id,
        #                     course=get_course.course)

        return redirect('KCHS:course_assessment_result')
